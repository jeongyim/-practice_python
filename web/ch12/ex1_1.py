from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook
wb = Workbook()
ws=wb.active
ws.append(["이름"])
br=webdriver.Chrome()
names =["유재석","송지효","김종국","하하","전소민","지석진"] 
for name in names:
    br.get('https://www.w3schools.com/tags/tryit.asp?filename=tryhtml_input_test')
    br.switch_to.frame('iframeResult')
    el1 = br.find_element(By.XPATH,'//*[@id="fname"]')
    el1.send_keys(name[1:])#name 2번째 문자 부터 끝까지 가져온다(이름부분)
    el2=br.find_element(By.XPATH,'//*[@id="lname"]')
    el2.send_keys(name[0])
    el3=br.find_element(By.XPATH,'/html/body/form/input[3]')
    time.sleep(3)
    el3.click()
    # time.sleep(2)
    # el4=br.find_element(By.XPATH,'/html/body/div[1]')
     #엑셀 파일 생성
    ws.append([name])
wb.save("webPeople.xlsx") ; print("첫번째 엑셀파일 생성완료") ; br.quit()

from openpyxl import load_workbook
br = webdriver.Chrome()
wb = load_workbook("webPeople.xlsx")
ws1=wb.create_sheet()
ws1.append(["두번째"])
for row in ws.iter_rows(min_row=2):
    name = row[0].value
    br.get('https://www.w3schools.com/tags/tryit.asp?filename=tryhtml_input_test')
    br.switch_to.frame('iframeResult')
    el1 = br.find_element(By.XPATH,'//*[@id="fname"]')
    el1.send_keys(name[1:])#name 2번째 문자 부터 끝까지 가져온다(이름부분)
    el2=br.find_element(By.XPATH,'//*[@id="lname"]')
    el2.send_keys(name[0])
    el3=br.find_element(By.XPATH,'/html/body/form/input[3]')
    time.sleep(3)
#print(el1.text)
    el3.click()
    ws1.append([name])
wb.save("webPeople.xlsx")
br.quit()
print("두번째 시트 생성 완료")
