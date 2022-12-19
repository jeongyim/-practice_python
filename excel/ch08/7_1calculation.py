#7_1calculation.py  날짜정보와 계산 과련 기능 실습해보기
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

import datetime #날짜정보를 가져오기 위해 임포트 함
ws["A1"] =datetime.datetime.today()#오늘 날짜 정보를 A1셀에 넣는다
ws.column_dimensions["A"].width=20
ws["A2"] = "=SUM(11,12,13)"#합
ws["A3"] ="=AVERAGE(11,12,13)"#평균
#정렬하기
from openpyxl.styles import Alignment #정렬을 위해 임포트
ws["B1"] = 10
ws["B2"] = 20
ws["B3"] = 30
ws["A4"].alignment = Alignment(horizontal="right") 
ws["A4"]="B1~B3 합계"
ws["B4"] = "=SUM(B1:B3)"
ws["A5"].alignment = Alignment(horizontal="right") 
ws["A5"] ="B1~B3 평균"
ws["B5"] ="=AVERAGE(B1:B3)"
wb.save("sample_calculation.xlsx") ; wb.close()