#1_1search.py

from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active
#국어 과목 점수가 80점 이상인 학생 찾기
for row in ws.iter_rows(min_row=2):
    #번호, 국어, 영어, 수학
    if int(row[1].value) >=80:#국어 점수가 80점 이상이면
        print(row[0].value,"번은 국어 80점 이상입니다.")

#영어과목명을 찾아서 컴퓨터 과목으로 변경
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value="컴퓨터"
wb.save("sample_change.xlsx")
wb.close()
