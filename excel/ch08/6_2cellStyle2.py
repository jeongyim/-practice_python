from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active
#셀에 배경색, font, 정렬을  위해 임포트함
from openpyxl.styles import PatternFill, Font, Alignment
#각 셀 가운데 정렬 후 
#90점 이상인 셀에 대해서는 초록색 적용
for row in ws.rows:    
    for cell in row:
        #각 cell에 대해서 정렬
        #horizontal:가로기준, vertical:세로기준 #center, left,right,bottom
        cell.alignment = Alignment(horizontal="center", vertical="center")
#번호열은 제외하고 점수가 90점 이상인 셀을 초록색으로 변경하기
    if cell.column == 1 :#A 번호열은 제외
        continue
        #셀의 자료형이 정수이고 90이상이면
    if isinstance(cell.value, int) and cell.value>=90:
            cell.fill = PatternFill(fgColor="00ff00", fill_type="solid")
            cell.font=Font(color="ff0000")#폰트 색상 변경
#틀 고정
ws.freeze_panes="B2"#B2를 기준으로 틀 고정
wb.save("sample_style2.xlsx") ; wb.close()
