#8_1merge.py
from openpyxl import Workbook
wb=Workbook()
ws=wb.active

ws.merge_cells("B2:D3")#B2부터 D3까지 병합
ws["B2"]="merged cell"
from openpyxl.styles import Alignment
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

wb.save("sample_merge.xlsx"); wb.close()
