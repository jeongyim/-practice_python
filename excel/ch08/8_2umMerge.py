#8_2unMerge.py
from openpyxl import load_workbook
wb=load_workbook("sample_merge.xlsx")
ws=wb.active
#B2:D3까지 병합된 셀 해제
ws.unmerge_cells("B2:D3")
ws.column_dimensions["B"].width = 20
wb.save("sample_unmerge.xlsx"); 
wb.close()
