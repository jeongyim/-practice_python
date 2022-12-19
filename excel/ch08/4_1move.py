
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

#3열에 컴퓨터 과목을 삽입하기 위하여 
#영어, 수학 과목을 한 칸 옆으로 옮김
ws.move_range("C1:D5",rows=0,cols=1)
ws["C1"].value="컴퓨터"
#국어과목명과 점수를 번호 아래로 이동
ws.move_range("B1:B5",rows=5,cols=-1)
wb.save("sample_move.xlsx")
wb.close()
