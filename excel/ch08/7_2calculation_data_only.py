#7_2calculation_data_only.py
# from openpyxl import load_workbook
# wb=load_workbook("sample_calculation.xlsx")
# ws=wb.active
# rows=ws[1:ws.max_row]
# # #수식 그대로 가져옴
# for row in rows:
#     for cell in row:
#         print(cell.value, end=" ")
#     print()

from openpyxl import load_workbook
wb=load_workbook("sample_calculation.xlsx", data_only=True)
ws=wb.active
rows=ws[1:ws.max_row]
# #수식이 아닌 실제 데이터를 그대로 가져옴
#evaluate(엑셀에서 직접저장) 되지 않은 상태의 데이터는None 라고 표시
for row in rows:
    for cell in row:
        print(cell.value, end=" ")
    print()
wb.close()
