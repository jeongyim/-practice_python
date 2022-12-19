#6_1cellStyle.py
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active
a1=ws["A1"]#번호
b1=ws["B1"]#국어
c1=ws["C1"]#영어
d1=ws["D1"]#수학
#번호의 셀 너비 좁히기
#A열의 너비를 5로 설정
ws.column_dimensions["A"].width=5
#1행의 높이를 50으로 설정
ws.row_dimensions[1].height=50
#글자색상 변경 등 스타일 적용을 위해 임포트
from openpyxl.styles import Font
#글자를 빨강,기울임,두껍게
a1.font =Font(color="ff0000", italic=True,bold=True)
#핑크보라색,Arial체, 취소선
b1.font=Font(color="cc33ff",name="Arial",strike=True)
#파란색, 크기20,밑줄 적용
c1.font=Font(color="0000ff",size=20, underline="single")
#셀에 테두리를 적용하기 위해서 Border와 Side임포트
from openpyxl.styles import Border, Side
#얇은 테두리 선 적용
multy_border=Border(left=Side(style="thin"),right=Side(style="dotted"),\
    top=Side(style="thick"),bottom=Side(style="double"))
b1.border = multy_border
c1.border = multy_border
d1.border = multy_border

wb.save("sample_style.xlsx")
wb.close()
