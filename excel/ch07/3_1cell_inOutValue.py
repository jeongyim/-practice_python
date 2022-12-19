from openpyxl import Workbook
wb=Workbook()
ws=wb.active#활성화 되어있는 sheet활용
#<셀에 값 입력>
ws["A1"]=1
ws["A2"]=2
ws["A3"]=3

ws["B1"]=4
ws["B2"]=5
ws["B3"]=6
#  행열번호를 이용해서 값 넣어 보기
#colum =열이름 A대신 1, B대신 2, C대신 3,.......으로 표현
#row = 행번호 1,2,3,표현
ws.cell(column=3, row=1, value=7)#ws["C1"]=7과 같음
ws.cell(column=3, row=2, value=8)#ws["C2"]=8과 같음

print(ws["A1"])
print(ws["A1"].value)
print(ws["A9"].value)

print(ws.cell(column=1, row=1).value)
print(ws.cell(column=2, row=1).value)
c = ws.cell(column=3, row=3, value=9)
print(c.value)
wb.save("sample.xlsx")
wb.close()

