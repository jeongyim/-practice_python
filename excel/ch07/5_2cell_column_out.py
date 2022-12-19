#5_2cell_column_out.py
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

col_A=ws["A"] #번호 컬럼을 가져옴
print(col_A) #셀 객체정보를 가져옴
#실행결과 <Cell ＇Sheet＇.A1>, <Cell ＇Sheet＇.A2>,
#  <Cell ＇Sheet＇.A3>, <Cell ＇Sheet＇.A4>, <Cell ＇Sheet＇.A5>)

#컬럼의 값 가져오기
#print(col_A.value)#에러 발생
for col in col_A:
    print(col.value) #객체를 하나씩 가져와서 값을 출력함

#여러 개의 컬럼을 가져옴
cols = ws["A:D"]#A컬럼부터 D컬럼까지의 값을 가져옴
for col in cols:
    for cell in col:
        print(cell.value,end=" ")
    print()
