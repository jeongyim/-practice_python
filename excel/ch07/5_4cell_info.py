from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active
#셀의 위치정보 가져오기
from openpyxl.utils.cell import coordinate_from_string
rows = ws[1:ws.max_row]
for row in rows:
    for cell in row:
        print(cell.coordinate, end=" ")
    print(    )
#셀의 위치정보를 컬럼과 row정보로 각각 분리해서 가져오기
for row in rows:
    for cell in row:
        xy=coordinate_from_string(cell.coordinate)
        print(xy, end=" ")
    print()
#컬럼과 row정보를 이용해서 셀 위치 나타내기
for row in rows:
    for cell in row:
        xy=coordinate_from_string(cell.coordinate)
        print(xy[0], end="")#컬럼정보
        print(xy[1],end=" ")#row정보
    print()
