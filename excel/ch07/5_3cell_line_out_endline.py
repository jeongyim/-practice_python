#5_3cell_line_out.py
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active

#첫번째 제목 줄 가져오기
row_1 = ws[1]
print(row_1)#첫번째 줄 객체 정보를 가져옴
#실행결과:(<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>, <Cell 'Sheet'.D1>)
for cell in row_1:
    print(cell.value,end=" ")
print()
#여러 줄 가져 오기
rows = ws[2:ws.max_row] #두번째 줄부터 5번째 줄까지 가져옴

for row in rows:
    for cell in row:
        #print(cell.value, end=" ")
        print(str(cell.value).rjust(4), end=" ")#4만큼 공간확보후 오른쪽정렬
    print()

