
#6_2cell_iter.py
from openpyxl import load_workbook
wb=load_workbook("sample.xlsx")
ws=wb.active
#전체 row정보 가져오기
print(ws.iter_rows())#알수없는 정보 출력:<generator object Worksheet._cells_by_row at 0x0000024B936E12A0>

# 각 row 정보를 출력
for row in ws.iter_rows():
    print(row)#하나의 줄씩 출력
for row in ws.iter_rows():
    print(row[1].value)#각 줄 중 두번째 값만 출력 할 것이다.
# 위와 같이 column에도 접근할 수 있다.
for col in ws.iter_cols():
    print(col)
for col in ws.iter_cols():
    print(col[0].value)
#범위를 지정해서 가져오기-앞코드에 이어서 코딩
print("*"*30)
for row in ws.iter_rows(min_row=1, max_row=3):#1번줄부터 3번줄까지 값 가져오기
    print(row[2].value)
#1번째 줄부터 3번째줄까지, 2번째 열부터 3번째 열까지 끊어서 가져옴
for row in ws.iter_rows(min_row=1, max_row=3,min_col=2,max_col=3):
    print(row[0].value,row[1].value)
    #print(row)

for col in ws.iter_cols(min_row=1, max_row=3,min_col=2,max_col=3):
    print(col)
   # print(col[0].value, col[1].value)

#min값을 지정하지 않으면 1로 시작
#max값을 지정하지 않으면 최대값으로 인식 함으로 항상 값을 지정하지 않아도 된다.
for row in ws.iter_rows(min_row=2):#두번째 줄부터 모든 데이터를 가지로 오겠다.
    print(row)
    print(row[1].value)
