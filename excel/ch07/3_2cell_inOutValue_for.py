from openpyxl import Workbook
wb = Workbook()
ws =wb.active
sum = 0
for x in range(1,6):
    for y in range(1,11):
        sum +=1
        ws.cell(column=x, row=y, value=sum)
sum = 0
for x in range(12,17):
    for y in range(1,11):
        sum +=1
        ws.cell(row=x, column=y, value=sum)

wb.save("sample.xlsx")
wb.close