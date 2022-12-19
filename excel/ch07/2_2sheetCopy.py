#2_2sheetCopy.py
from openpyxl import Workbook
wb=Workbook()
ws1=wb.create_sheet("1pySheet")
ws2=wb.create_sheet("2pySheet")
ws3=wb.create_sheet("3pySheet")
#sheet 복사
ws1["A1"]="test"#1pySheet A1셀에 "test"라는 데이터가 들어감
wsc=wb.copy_worksheet(ws1)#ws1 sheet를 복사해서 wsc sheet라는 복사본을 만든다.
wsc.title="copyedSheet"

wb.save("sample.xlsx")
wb.close()
