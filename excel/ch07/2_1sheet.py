#2_1sheet.py
from openpyxl import Workbook
wb=Workbook()
#wb.active : 활성화된 시트 가져옴
ws=wb.create_sheet()#새로운 sheet 기본이름으로 생성
ws.title = "MyPySheet" #sheet이름 변경

ws.sheet_properties.tabColor = "ccccff" #RGB형태로 값을 넣어주면 탭 색상 변경
ws1=wb.create_sheet("OkPySheet")#주어진 이름으로 sheet생성
ws1.sheet_properties.tabColor = "0000ff"
#Sheet:0번지,MyPySheet:1번지,OkPySheet:2번지
ws2 = wb.create_sheet("NoPySheet",2) #2번지에 sheet생성
li_ws = wb["MyPySheet"]#워크북으로 접근
li_ws.sheet_properties.tabColor = "ff0000"
wb["Sheet"].title = "Change"
print(wb.sheetnames)#모든 sheet이름 출력
name = wb.sheetnames
print(name[0]) #0번지 시트이름 출력
wb.save("sample.xlsx")
#열려있는 sample.xlsx가 있으면 실행시 오류가 남으로 sample.xlsx창을 닫고 실행한다.
wb.close()