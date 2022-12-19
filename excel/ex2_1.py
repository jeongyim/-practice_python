# from openpyxl import Workbook
# wb=Workbook()
# ws=wb.active
# ws["A1"]="사과"
# ws["B1"]="바나나"
# ws["C1"]="오렌지"

# sum = 0
# for x in range(1,2):
#     for y in range(2,7):
#         sum +=1
#         ws.cell(column=x, row=y, value=sum)

# sum = 5
# for x in range(2,3):
#     for y in range(2,7):
#         sum +=1
#         ws.cell(column=x, row=y, value=sum)

# sum = 10
# for x in range(3,4):
#     for y in range(2,7):
#         sum +=1
#         ws.cell(column=x, row=y, value=sum)

# wb.save("text.xlsx") 
# wb.close() 

from openpyxl import Workbook
wb=Workbook()
ws=wb.active

ws.append(["사과","바나나","오렌지"])
sum=0
for x in range(1,4):
    for y in range(2,7):
        sum=sum+1
        ws.cell(column=x,row=y,value=sum)
wb.save("test.xlsx")
wb.close()

