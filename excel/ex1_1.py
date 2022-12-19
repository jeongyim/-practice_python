# from openpyxl import Workbook
# wb=Workbook()
# ws=wb.active
# ws2 = wb.create_sheet("NoPySheet")
# ws2.title = "apple"
# ws2.sheet_properties.tabColor = "ff0000"
# ws3 = wb.create_sheet("orange",2)
# ws3.sheet_properties.tabColor = "ffff00"
# ws3["A1"]="비타민 c가 많음"
# wsc=wb.copy_worksheet(ws3)
# wsc.title="orange2"
# print(wb.sheetnames)
# name = wb.sheetnames
# print(name[2])
# wb.save("sample.xlsx")
# wb.close()


from openpyxl import Workbook
wb=Workbook()
ws=wb.active
ws.title="fruit"

ws1=wb.create_sheet("apple")
ws1.sheet_properties.tabColor="ff0000"

ws2=wb.create_sheet("orange",1)#두번째 sheet는 1번지
ws2.sheet_properties.tabColor="ffff66"
ws2["A1"]="비타민 c가 많음"

wsc=wb.copy_worksheet(ws2)
wsc.title="orange2"

print(wb.sheetnames)
print(wb.sheetnames[1])

wb.save("test.xlsx")
wb.close()
