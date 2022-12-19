# from openpyxl import load_workbook 
# wb=load_workbook("text.xlsx")
# ws=wb.active 

# rows = ws[1:6]
# for row in rows:
#     for cell in row:
#         print(cell.value, end=" ")       
#     print()

# print("-------------------")

# for cell in ws[1]:
#     print(cell.value,end=" ")
# print()

# print("-------------------")

# rows = ws[2:6]
# for row in rows:
#     for cell in row:
#         print(cell.value, end=" ")       
#     print()


#실습 3
from openpyxl import load_workbook
wb=load_workbook("test.xlsx")
ws=wb.active
for x in range(1, ws.max_row+1):
    for y in range(1,ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()
print("<과일 이름만 출력>")
for row in ws.iter_rows(max_row =1):
     for cell in row:
         print(cell.value, end=" ")
print()
print("<숫자값만 출력>")
for row in ws.iter_rows(min_row=2):
     for cell in row:
        print(cell.value, end=" ")
     print()
print("<B컬럼에 있는 값만 출력>")
for row in ws.iter_rows():
    print(row[1].value)
