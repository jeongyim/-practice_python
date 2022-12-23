from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage
from account import*
wb=load_workbook("sendPeople.xlsx")
ws=wb.active
with smtplib.SMTP("smtp.gmail.com",587)as smtp:
    smtp.ehlo()
    smtp.starttls()
    smtp.login(EMAIL_ADDRESS,EMAIL_PASSWORD)
    for row in ws.iter_rows(min_row=3):
        if row[2].value:#이메일 주소란에 값이 존재하면 아래문장을 실행
            msg=EmailMessage()
            msg["Subject"]="excel파일을 이용 메시지 발송"
            msg["From"]=EMAIL_ADDRESS
            msg["To"]=row[2].value
            content="{}님 엑셀 파일에 있는 메일주소로 메일을 보냈습니다.".format(row[1].value)
            msg.set_content(content)
            smtp.send_message(msg)
            print(row[1].value,"님에게 메일 발송완료")
